"""
GTFS Rail Simulator — Metro, Marmaray, Tramvay, Füniküler
Outputs: public/data/rail_sim.json
         public/data/turkey_overlays/kent_lokantasi.geojson
         public/data/turkey_overlays/sosyal_tesisler.geojson
"""
import csv, json, os, math
from collections import defaultdict
import openpyxl

BASE   = os.path.join(os.path.dirname(__file__), '..', 'public', 'gtfs and more')
DATAS  = os.path.join(os.path.dirname(__file__), '..', 'public', 'datas')
OUT    = os.path.join(os.path.dirname(__file__), '..', 'public', 'data', 'rail_sim.json')
OVL    = os.path.join(os.path.dirname(__file__), '..', 'public', 'data', 'turkey_overlays')

RAIL_TYPES = {'0', '1', '4', '7'}  # tram, metro/suburban, ferry, funicular

ROUTE_COLORS = {
    'M1A': [230,0,18], 'M1B': [230,0,18],
    'M2':  [230,0,18], 'M2A': [230,0,18],
    'M3':  [112,34,131], 'M3A': [112,34,131],
    'M4':  [247,148,29],
    'M5':  [61,181,74],
    'M6':  [130,195,65],
    'M7':  [181,160,52],
    'M8':  [250,210,1],
    'M9':  [27,117,187],
    'T1':  [0,163,216],
    'T3':  [141,198,63],
    'T4':  [27,117,187],
    'F1':  [125,78,45], 'F2': [125,78,45], 'F3': [125,78,45],
    'Marmaray':  [26,58,108],
    'Marmaray1': [26,58,108],
    'Marmaray2': [26,58,108],
}

DEFAULT_HEADWAYS = {
    'M1A':240,'M1B':240,'M2':240,'M2A':240,
    'M3':420,'M3A':420,'M4':300,'M5':300,
    'M6':600,'M7':360,'M8':360,'M9':480,
    'T1':120,'T3':300,'T4':180,
    'F1':300,'F2':300,'F3':300,
    'Marmaray':480,'Marmaray1':480,'Marmaray2':480,
    # Vapurlar — saatte 1-4 sefer
    '_ferry_default': 1800,  # 30 dk varsayılan
}

def fix_str(s):
    # GTFS dosyası ISO-8859-9 (Türkçe Latin-5) ile encode edilmiş,
    # utf-8-sig ile okunduğunda ? karakterleri oluşmuş.
    # Doğru okuma: ham baytları iso-8859-9 ile decode et.
    return s  # artık doğrudan okuyoruz, encode sonrası fix gerekmez

FERRY_COLOR = [14, 116, 144]   # cyan-700

def get_kind(short, route_type='1'):
    if route_type == '4': return 'ferry'
    if 'Marmaray' in short: return 'marmaray'
    if short.startswith('M'):  return 'metro'
    if short.startswith('T'):  return 'tram'
    if short.startswith('F'):  return 'funicular'
    return 'metro'

def parse_time(s):
    p = s.strip().split(':')
    if len(p) != 3: return -1
    return int(p[0])*3600 + int(p[1])*60 + int(p[2])

# ── 1. Routes ──────────────────────────────────────────────────────────────────
print("Reading routes…")
rail_routes = {}
with open(f'{BASE}/routes.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        if r.get('route_type') not in RAIL_TYPES: continue
        rid   = r['route_id']
        rtype = r.get('route_type','1')
        short = fix_str(r.get('route_short_name','').strip())
        kind  = get_kind(short, rtype)
        # Ferry rengi sabit cyan, diğerleri harita veya varsayılan
        if kind == 'ferry':
            color = FERRY_COLOR
        else:
            color = ROUTE_COLORS.get(short, [100,100,100])
            gc = r.get('route_color','').strip()
            if gc and len(gc) == 6:
                try: color = [int(gc[i:i+2],16) for i in (0,2,4)]
                except: pass
        rail_routes[rid] = dict(short=short, long=fix_str(r.get('route_long_name','').strip()),
                                color=color, kind=kind)
print(f"  {len(rail_routes)} rail routes")

# ── 2. Calendar — hangi servis hangi gün çalışıyor ─────────────────────────────
# Simülasyonun tarih kavramı yok; tek bir temsili HAFTA İÇİ günü (Cuma) tarifesini
# "pişiriyoruz". Böylece sadece o gün sefer yapan hatlar üretiliyor — hafta sonu /
# sezonluk / tur hatları (örn. Boğaz turları) gereksiz yere aktif görünmüyor.
TARGET_DAY = 4  # 0=Pzt … 4=Cuma, 6=Paz
_DAY_COLS = ['monday','tuesday','wednesday','thursday','friday','saturday','sunday']
print("Reading calendar…")
service_active = {}  # service_id -> bool (TARGET_DAY'de çalışıyor mu)
try:
    with open(f'{BASE}/calendar.csv', encoding='iso-8859-9') as f:
        for r in csv.DictReader(f):
            service_active[r['service_id']] = (r.get(_DAY_COLS[TARGET_DAY], '0') == '1')
except FileNotFoundError:
    print("  calendar.csv yok — tüm servisler aktif sayılacak")
def runs_today(sid):
    # Bilinmeyen service_id → dahil et (veri kaybı olmasın)
    return service_active.get(sid, True)
print(f"  {len(service_active)} servis, {sum(service_active.values())} tanesi hedef günde aktif")

# ── 3. Trips — (route, direction) başına TÜM seferler, hedef güne göre filtreli ─
print("Reading trips…")
trips_by_key = defaultdict(list)  # (route_id, dir) -> [{trip_id, shape_id, headsign}]
with open(f'{BASE}/trips.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        rid = r.get('route_id')
        if rid not in rail_routes: continue
        sid = r.get('shape_id','').strip()
        if not sid: continue
        if not runs_today(r.get('service_id','')): continue  # bugün çalışmıyorsa atla
        key = (rid, r.get('direction_id','0'))
        trips_by_key[key].append(dict(trip_id=r['trip_id'], shape_id=sid,
                                      headsign=fix_str(r.get('trip_headsign','').strip())))
needed_trips = {t['trip_id'] for lst in trips_by_key.values() for t in lst}
print(f"  {len(trips_by_key)} route+yön grubu, {len(needed_trips)} hedef-gün seferi")

# ── 4. Shapes ─────────────────────────────────────────────────────────────────
print("Reading shapes…")
needed_shapes = {t['shape_id'] for lst in trips_by_key.values() for t in lst}
shapes = defaultdict(list)
with open(f'{BASE}/shapes.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        sid = r.get('shape_id','').strip()
        if sid not in needed_shapes: continue
        shapes[sid].append((int(r['shape_pt_sequence']),
                            float(r['shape_pt_lon']), float(r['shape_pt_lat'])))
for sid in shapes:
    shapes[sid].sort()
    shapes[sid] = [[lon,lat] for _,lon,lat in shapes[sid]]
print(f"  {len(shapes)} shapes loaded")

# ── 5. Stops (isimler) ──────────────────────────────────────────────────────────
print("Reading stops…")
stop_names = {}  # stop_id -> name
with open(f'{BASE}/stops.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        stop_names[r['stop_id']] = fix_str((r.get('stop_name') or '').strip())
print(f"  {len(stop_names)} stops")

# ── 6. Stop times — hedef-gün seferlerinin gerçek kalkış saatleri ──────────────
print("Reading stop_times…")
trip_times = defaultdict(list)  # trip_id -> [(seq, dep_sec, stop_name)]
with open(f'{BASE}/stop_times.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        tid = r.get('trip_id','')
        if tid not in needed_trips: continue
        dep = parse_time(r.get('departure_time',''))
        if dep < 0: continue
        name = stop_names.get(r.get('stop_id',''), '')
        trip_times[tid].append((int(r.get('stop_sequence',0)), dep, name))
for tid in trip_times:
    trip_times[tid].sort()
print(f"  stop_times for {len(trip_times)} trips")

# ── 7. Frequencies — gerçek servis pencereleri (start/end/headway) ─────────────
print("Reading frequencies…")
trip_freq = defaultdict(list)  # trip_id -> [(start_sec, end_sec, headway_secs)]
with open(f'{BASE}/frequencies.csv', encoding='iso-8859-9') as f:
    for r in csv.DictReader(f):
        tid = r.get('trip_id','')
        if tid not in needed_trips: continue
        st = parse_time(r.get('start_time',''))
        et = parse_time(r.get('end_time',''))
        hw = int(r.get('headway_secs', 0) or 0)
        if hw > 0 and st >= 0 and et > st:
            trip_freq[tid].append((st, et, hw))
print(f"  frequency pencereleri: {sum(len(v) for v in trip_freq.values())} satır, {len(trip_freq)} sefer")

# ── 8. Build compact format ───────────────────────────────────────────────────
# Format: routes store path+duration once; trips store only (route_key, t0)
# Sefer saatleri ARTIK gerçek tarifeden geliyor:
#   • frequency penceresi varsa → pencere içinde headway aralıklarla üret
#   • yoksa → seferin GERÇEK kalkış saatini kullan (tek tek)
print("Building rail_sim.json (compact format)…")
route_defs = {}  # route_key -> {name, headsign, color, kind, path, duration_secs}
trips_out  = []  # [{rk, t0}]

def downsample(path, max_pts=200):
    """Keep at most max_pts evenly spaced points."""
    n = len(path)
    if n <= max_pts: return path
    step = n / max_pts
    return [path[int(i*step)] for i in range(max_pts)]

for (rid, dir_), lst in trips_by_key.items():
    route = rail_routes[rid]
    short = route['short']
    color = route['color']
    kind  = route['kind']

    # Temsili sefer: geçerli shape + en az 2 duraklı ilk sefer
    rep = None
    for t in lst:
        path = shapes.get(t['shape_id'], [])
        stops = trip_times.get(t['trip_id'], [])
        if len(path) >= 2 and len(stops) >= 2 and stops[-1][1] > stops[0][1]:
            rep = t; break
    if rep is None: continue

    path  = shapes[rep['shape_id']]
    stops = trip_times[rep['trip_id']]
    rt0   = stops[0][1]
    duration = stops[-1][1] - rt0

    # Durak listesi: [{name, elapsed_secs}] — t0'dan itibaren geçen süre
    stop_list = [
        {'name': name, 'elapsed_secs': dep - rt0}
        for _, dep, name in stops
        if name  # isimsiz durakları atla
    ]

    rk = f"{short}|{dir_}"
    route_defs[rk] = dict(
        name=short,
        headsign=rep['headsign'],
        color=color,
        kind=kind,
        path=downsample(path, 150),
        duration_secs=duration,
        stops=stop_list,
    )

    # Kalkış saatleri — gerçek tarifeden
    freq_windows = [w for t in lst for w in trip_freq.get(t['trip_id'], [])]
    t0_set = set()
    if freq_windows:
        # Frequency tabanlı hat (metro/marmaray/tram vb.): pencere içinde üret
        for st, et, hw in freq_windows:
            t = st
            while t < et:
                t0_set.add(t)
                t += hw
    else:
        # Tarifeli hat (çoğu vapur): her seferin gerçek kalkış saati
        for t in lst:
            stt = trip_times.get(t['trip_id'])
            if stt:
                t0_set.add(stt[0][1])

    for t0 in sorted(t0_set):
        trips_out.append({'rk': rk, 't0': t0})

total_trips = len(trips_out)
print(f"  {len(route_defs)} route definitions, {total_trips} trips")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(dict(routes=route_defs, trips=trips_out), f, ensure_ascii=False)
print(f"  Written: {OUT}  ({os.path.getsize(OUT)//1024} KB)")

# ── 7. Kent Lokantası (xlsx) ──────────────────────────────────────────────────
print("Converting kent-lokantalar…")
def auto_coord(v, is_lat=True):
    """Convert integer coord (unknown scale) to decimal degrees."""
    if v is None: return None
    iv = abs(int(float(v)))
    if iv == 0: return None
    digits = len(str(iv))
    scale = 10 ** (digits - 2)
    result = float(v) / scale
    rng = (38, 43) if is_lat else (26, 32)
    if rng[0] <= result <= rng[1]:
        return result
    # try ±1 scale
    for s in [scale*10, scale//10]:
        if s == 0: continue
        r = float(v) / s
        if rng[0] <= r <= rng[1]: return r
    return None

features = []
wb = openpyxl.load_workbook(f'{BASE}/xlsx/kent-lokantalar-lokasyon.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    name, lat_raw, lon_raw = row[0], row[1], row[2]
    addr = row[3] if len(row) > 3 else ''
    lat = auto_coord(lat_raw, is_lat=True)
    lon = auto_coord(lon_raw, is_lat=False)
    if lat is None or lon is None: continue
    features.append({'type':'Feature',
        'properties':{'name': name or '', 'address': addr or ''},
        'geometry':{'type':'Point','coordinates':[lon, lat]}})

out_path = os.path.join(OVL, 'kent_lokantasi.geojson')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'type':'FeatureCollection','features':features}, f, ensure_ascii=False)
print(f"  {len(features)} kent lokantalari -> {out_path}")

# ── 8. Sosyal Tesisler (xlsx) ─────────────────────────────────────────────────
print("Converting sosyal-tesisler…")
features = []
wb = openpyxl.load_workbook(f'{BASE}/xlsx/ibb-sosyal-tesis-konumlar.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    name, lat_raw, lon_raw = row[0], row[1], row[2]
    addr = row[3] if len(row) > 3 else ''
    lat = auto_coord(lat_raw, is_lat=True)
    lon = auto_coord(lon_raw, is_lat=False)
    if lat is None or lon is None: continue
    features.append({'type':'Feature',
        'properties':{'name': name or '', 'address': addr or ''},
        'geometry':{'type':'Point','coordinates':[lon, lat]}})

out_path = os.path.join(OVL, 'sosyal_tesisler.geojson')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'type':'FeatureCollection','features':features}, f, ensure_ascii=False)
print(f"  {len(features)} sosyal tesis -> {out_path}")

print("\nAll done!")
